import serial
import openpyxl # 外部ライブラリ　pip install openpyxl
import os
import time

pathinput = "./"+input("ファイル名:")+".xlsx"
#pathinput = "./test1.xlsx"
print(pathinput)
in_book = openpyxl.load_workbook(pathinput)           #数式を取得
in_bookdata = openpyxl.load_workbook(pathinput, data_only=True)           #数式の値を取得
pathsheet = input("シート名:")
print(pathsheet)
in_wb0 = in_book[ pathsheet ]
in_wb0_data = in_bookdata[ pathsheet ]
lastrows0 = in_wb0.max_row
lastcols0 = in_wb0.max_column
print(pathsheet,lastrows0,lastcols0)

# シリアルポートの設定
COMNo = 'COM'+str(in_wb0.cell(2,3).value)
serial_port = COMNo  # 使用するシリアルポート名を指定します
baud_rate = in_wb0.cell(3,3).value

ComNum = in_wb0_data.cell(8,3).value

for n in range(1,ComNum):
    # 送信するコマンド（バイト配列）
    line=input(" 送信行No?:")
    for i in range(0,lastrows0-1):
        if str(line) == str(in_wb0_data.cell(i+1,5).value):
            # データを読み取り、バイト配列に変換
            command = bytearray()
            for j in range(0,13):
                cell_value = in_wb0_data.cell(i+1,6+j).value
                if isinstance(cell_value,str):
                    byte_value = int(cell_value,16)
                elif isinstance(cell_value, int):
                    byte_value = cell_value
                else:
                    raise ValueError(f"Unexpected cell value: {cell_value}")
                command.append(byte_value)
            print("Command to send:", command)
            
            # シリアルポートを開く
            with serial.Serial(serial_port, baud_rate, timeout=1) as ser:
                #コマンドを送信
                ser.write(command)

                time.sleep(in_wb0_data.cell(i+1,19).value / 1000)
                
                #レスポンス
                response = ser.read(13)  # 応答のバイト数を指定します
                print("Response:", response)
                res = ["FF","FF","FF","FF","FF","FF","FF","FF","FF","FF","FF"]
                n = 0
                for byte in response:
                    res[n] = format(int(byte) & 0xFF, '02X')
                    in_wb0.cell(i+1,22+n).value = res[n]
                    n += 1
        
    in_book.save("output.xlsx")
    #time.sleep(2)
    in_book.close
    os.startfile("output.xlsx")
    input("enterキー:")
    time.sleep(2)
    os.close

    #EXCELの再読み込み
    print(pathinput)
    in_book = openpyxl.load_workbook(pathinput)           #数式を取得
    in_bookdata = openpyxl.load_workbook(pathinput, data_only=True)           #数式の値を取得
    print(pathsheet)
    in_wb0 = in_book[ pathsheet ]
    in_wb0_data = in_bookdata[ pathsheet ]
    lastrows0 = in_wb0.max_row
    lastcols0 = in_wb0.max_column
    print(pathsheet,lastrows0,lastcols0)